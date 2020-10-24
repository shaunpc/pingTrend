# PingTrend - COLLECT-EXCEL.PY
#   to record results from PING command
#   both min/max/ave for each minute,
#   and the error types/rates
# and STORES in EXCEL file
#   .... in my GOOGLE-DRIVE folder, which gets syncs to cloud
#   .... and my AppSheet tool picks it up automatically...
#   https://www.appsheet.com/start/2aa910a9-b1bd-44e2-9caf-835da802c1bf#appName=UntitledApp-1995485

import subprocess
import datetime
import sys
import time
from counter import Counter
import os.path
import openpyxl as pyxl


# Executes a PING on the passed address, and returns either response time or the error message returned
def ping(host):
    error = None
    r_time = -1
    exec_ping = subprocess.run(["ping", "-n", "1" "-4", host], capture_output=True)
    if exec_ping.returncode == 0:
        lines = exec_ping.stdout.splitlines()
        if lines[2][:10] == b'Reply from':
            r_time = lines[2].split()[4][5:-2].decode('utf-8')
        else:
            error = "Error(0): {}".format(lines[2])
    else:
        error = "Error({}): {}".format(exec_ping.returncode, exec_ping.stdout)
    return float(r_time), error


def store_data(file, tab, values):
    wb = pyxl.load_workbook(file)
    for v in values:
        wb[tab].append(v)
    try:
        wb.save(datastore)
        wb.close()
    except IOError:
        print("IOERROR: Unable to store, preserving in memory: ")
        return values
    else:
        return []


# Main Routine.
if __name__ == '__main__':
    datastore = r'C:\Users\shaun\Google Drive\Data\pingtrend.xlsx'

    # if "CLEAN" is passed as param, and file exists then delete it
    if 'clean' in sys.argv and os.path.isfile(datastore):
        print("CLEAN: Removing existing datastore file: {}".format(datastore))
        os.remove(datastore)

    if not os.path.isfile(datastore):
        # create file if it doesn't exist
        print("CREATE: Creating datastore file: {}".format(datastore))
        workbook = pyxl.Workbook()
        ws_data = workbook['Sheet']
        ws_data.title = 'PingData'
        ws_data.append(['TimeStamp', 'Target', 'Min', 'Avg', 'Max', 'ErrorCount'])
        ws_error = workbook.create_sheet('PingErrors')
        ws_error.append(['TimeStamp', 'Target', 'Type', 'Message'])
        workbook.save(filename=datastore)
        workbook.close()

    ping_target = "www.google.com"

    # Start the counters
    counter = Counter("DATA")
    counter_err = Counter("ERROR")
    timestamp = datetime.datetime.now()
    last_min = timestamp.minute
    data_error = []
    data_pings = []
    print("Starting data collection at {}".format(timestamp))
    try:
        while True:
            time.sleep(0.25)
            timestamp = datetime.datetime.now()
            this_min = timestamp.minute
            rt, err_full = ping(ping_target)
            if err_full:
                print("ERROR: {}".format(err_full))
                err_type = "UNKNOWN"
                if "Request timed out" in err_full:
                    err_type = "TIMEOUT"
                if "Ping request could not find host" in err_full:
                    err_type = "NO HOST"
                data_error.append([timestamp, ping_target, err_type, err_full])
                data_error = store_data(datastore, 'PingErrors', data_error)
                counter_err.add(1)
            else:
                counter.add(rt)

            # Store the min/ave/max/error data each minute
            if this_min != last_min:
                print("LOG: {} {}/{}/{} (Err:{})".format(timestamp.strftime('%H:%M'), counter.min(),
                                                         counter.ave(), counter.max(), counter_err.count()))
                data_pings.append([timestamp, ping_target, counter.min(), counter.ave(),
                                   counter.max(), counter_err.count()])
                data_pings = store_data(datastore, 'PingData', data_pings)
                last_min = this_min
                counter.reset()
                counter_err.reset()
    except KeyboardInterrupt:
        print("Finishing data collection at {}".format(timestamp))
        print("LOG: {} {}/{}/{} (Err:{})".format(timestamp.strftime('%H:%M'), counter.min(),
                                                 counter.ave(), counter.max(), counter_err.count()))
        data_pings.append([timestamp, ping_target, counter.min(), counter.ave(), counter.max(), counter_err.count()])
        data_pings = store_data(datastore, 'PingData', data_pings)